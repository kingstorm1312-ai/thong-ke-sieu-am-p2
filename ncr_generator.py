import os
import time
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font 
from openpyxl.cell.text import InlineFont 
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import get_column_letter
import pandas as pd
import pythoncom
import win32com.client
import abc

class NCRGenerator:
    def __init__(self, template_path):
        self.template_path = template_path
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template not found: {self.template_path}")
        self._merge_cache = {}

    def _get_target_cell(self, sheet, coord):
        """
        Returns the actual cell to write to with Caching.
        """
        # Check Cache
        if coord in self._merge_cache:
            r, c = self._merge_cache[coord]
            return sheet.cell(row=r, column=c)

        cell = sheet[coord]
        if isinstance(cell, MergedCell):
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Cache the result
                    self._merge_cache[coord] = (merged_range.min_row, merged_range.min_col)
                    return sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
        
        # Determine if we should cache "Normal" cells? 
        # No, normal cells are fast. Merged check is the slow part.
        return cell

    def fill_template(self, data, output_path):
        """
        Fills the Excel template with data.
        data: dictionary containing keys corresponding to fields.
        output_path: path to save the filled Excel file.
        """
        wb = openpyxl.load_workbook(self.template_path)
        sheet = wb.active

        # Helper to set value safely
        def set_val(coord, val):
            if val is not None:
                target_cell = self._get_target_cell(sheet, coord)
                target_cell.value = val
                return target_cell
            return None

        # Helper to replace text safely
        def replace_in_cell(coord, old_text, new_text):
            target_cell = self._get_target_cell(sheet, coord)
            current_val = target_cell.value
            if current_val and isinstance(current_val, str):
                target_cell.value = current_val.replace(old_text, str(new_text))
            else:
                if not current_val:
                    target_cell.value = str(new_text)

        # 1. Fill Header Info
        
        # Date at A5
        if 'date_str' in data:
            set_val('A5', f"Ngày báo cáo: {data['date_str']}")

        # A7: Hợp đồng/ SO no :................
        if 'contract' in data:
            replace_in_cell('A7', '....................................................', f" {data['contract']}")
            cell_A7 = self._get_target_cell(sheet, 'A7')
            if cell_A7.value and "Hợp đồng" in str(cell_A7.value):
                 pass 
            val = cell_A7.value
            if val and str(val).strip().endswith(':'):
                 set_val('A7', f"{val} {data['contract']}")
            elif val and "..." in str(val):
                 set_val('A7', str(val).replace('...', f" {data['contract']}").replace('…', '')) 
            else:
                 set_val('A7', f"Hợp đồng/ SO no: {data['contract']}")

        # D7: Số lượng/ Quatity :……………
        if 'quantity' in data:
            cell_D7 = self._get_target_cell(sheet, 'D7')
            val = str(cell_D7.value) if cell_D7.value else ""
            if "..." in val or "…" in val:
                new_val = val.replace('…', '').replace('...', '') + f" {data['quantity']}"
                set_val('D7', new_val)
            else:
                set_val('D7', f"Số lượng/ Quantity: {data['quantity']}")

        # Helper to set value with Label and Auto-fit Font
        def set_val_with_label_fit(coord, label, val):
            if not val: return
            
            # Construct full text (Preserve Label)
            # Check if cell already has label? User implies 'giữ lại', 
            # simplest way is to re-construct it: "Label: Content"
            full_text = f"{label}: {val}"
            
            target_cell = self._get_target_cell(sheet, coord)
            target_cell.value = full_text
            
            # Font Resizing Logic
            # Baseline: assume standard width. If len > 45 -> reduce size.
            if len(full_text) > 40:
                new_size = 11 # Default apprx
                if len(full_text) > 80: new_size = 8
                elif len(full_text) > 60: new_size = 9
                elif len(full_text) > 45: new_size = 10
                
                # Apply new size (Preserving other attributes if possible)
                if target_cell.font:
                    f = target_cell.font
                    target_cell.font = Font(name=f.name, size=new_size, bold=f.b, italic=f.i, color=f.color)
                else:
                    target_cell.font = Font(size=new_size)

        # A6: Tên hàng (From G2)
        if 'item_name' in data:
            set_val_with_label_fit('A6', "Tên hàng", data['item_name'])
            
        # G7: Mã vật tư (From F2)
        if 'item_code' in data:
            set_val_with_label_fit('G7', "Mã vật tư", data['item_code'])

        # D6: Mã hàng
        if 'item_code' in data:
             set_val_with_label_fit('D6', "Mã hàng", data['item_code'])

        # D5: NCR No
        if 'ncr_no' in data:
             set_val('D5', f"NCR No: {data['ncr_no']}")

        # 2. Fill Defects Table
        # Rows 10+ are merged. Writing to Column A.
        curr_row = 10
        if 'defects' in data and isinstance(data['defects'], list):
            
            for idx, defect in enumerate(data['defects']):
                name = defect.get('name', '')
                qty = defect.get('qty', 0)
                rate = defect.get('rate', 0)
                
                # Format: "1. Name: Qty (Rate%)"
                # User wants Left Alignment and Bold Quantity (including Rate)
                target_cell = self._get_target_cell(sheet, f'A{curr_row}')
                
                # Create RichText: "1. Name: " (Normal) + "Qty (Rate%)" (Bold)
                # Note: Use InlineFont for TextBlock, NOT standard Font
                # Font(b=True) in InlineFont syntax
                
                prefix = f"{idx + 1}. {name}: "
                if rate > 0:
                    qty_str = f"{qty:,.0f} ({rate:.2f}%)"
                else:
                    qty_str = f"{qty:,.0f}"
                
                # [FIX] Use simple string for reliability
                full_text_def = f"{prefix}{qty_str}"
                
                target_cell.value = full_text_def
                target_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                curr_row += 1
        
        # 3. Fill Summary Table (Supplemental)
        if 'summary' in data and isinstance(data['summary'], list) and len(data['summary']) > 0:
            # Spacer
            curr_row += 1
            
            # Header
            header_cell = self._get_target_cell(sheet, f'A{curr_row}')
            header_cell.value = "TỔNG HỢP NGUYÊN NHÂN LỖI:"
            header_cell.font = Font(bold=True, underline='single')
            header_cell.alignment = Alignment(horizontal='left', vertical='center')
            curr_row += 1
            
            for s_item in data['summary']:
                s_name = s_item.get('name', '')
                s_qty = s_item.get('qty', 0)
                s_rate = s_item.get('rate', 0)
                
                # Format: Name: Qty (Rate%)
                # Format: Name: Qty (Rate%)
                t_cell = self._get_target_cell(sheet, f'A{curr_row}')
                
                # [FIX] Use simple string to avoid PDF conversion issues with RichText
                full_text_val = f"{s_name}: {s_qty:,.0f} ({s_rate:.2f}%)"
                print(f"✍️ [NCR Gen] Writing Summary Row {curr_row}: {full_text_val}")
                
                t_cell.value = full_text_val
                # [STYLE] Bold the entire cell for emphasis (Safe way)
                if t_cell.font:
                    f = t_cell.font
                    t_cell.font = Font(name=f.name, size=f.size, bold=True, italic=f.i, color=f.color)
                else:
                    t_cell.font = Font(bold=True) 
                
                t_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                curr_row += 1

        # 4. Save
        wb.save(output_path)
        return output_path

    def convert_to_pdf(self, excel_path, pdf_path):
        """
        Converts Excel file to PDF using win32com.
        """
        try:
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            excel.EnableEvents = False
            
            abs_excel_path = os.path.abspath(excel_path)
            abs_pdf_path = os.path.abspath(pdf_path)
            
            try:
                wb = excel.Workbooks.Open(abs_excel_path)
                # ExportAsFixedFormat 0 is xlTypePDF
                wb.ExportAsFixedFormat(0, abs_pdf_path)
                wb.Close(SaveChanges=False)
            except Exception as e:
                raise e
            finally:
                excel.Quit()
                
            return pdf_path
        except Exception as e:
            print(f"PDF Conversion failed: {e}")
            return None

    def generate(self, data, output_dir="temp_ncr"):
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        timestamp = int(time.time())
        excel_name = f"NCR_{timestamp}.xlsx"
        pdf_name = f"NCR_{timestamp}.pdf"
        
        excel_path = os.path.join(output_dir, excel_name)
        pdf_path = os.path.join(output_dir, pdf_name)
        
        # 1. Fill
        t_start = time.time()
        self.fill_template(data, excel_path)
        t_fill = time.time()
        print(f"⏱️ [NCR] Template Fill Time: {t_fill - t_start:.2f}s")
        
        # 2. Convert
        print("🔄 [NCR] Starting Excel -> PDF Conversion (Please wait)...")
        final_pdf = self.convert_to_pdf(excel_path, pdf_path)
        t_convert = time.time()
        print(f"⏱️ [NCR] PDF Conversion Time: {t_convert - t_fill:.2f}s")
        
        return final_pdf if final_pdf else excel_path # Return Excel if PDF fails

